import openpyxl

data_list = []

try:
    data_file = openpyxl.load_workbook("data.xlsx")
except FileNotFoundError:
    data_file = openpyxl.load_workbook("newdata.xlsx")

# fetch the active sheet
sheet = data_file.active

# iterate through each row and store the data in the data list
for row in sheet.iter_rows():
    row_data = []
    for eachcell in row:
        # print(eachcell.value)
        row_data.append(eachcell.value)
    data_list.append(row_data)

print(data_list)
